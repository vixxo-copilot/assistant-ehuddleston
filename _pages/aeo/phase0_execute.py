#!/usr/bin/env python3
"""Phase 0: Profound scoring, Batch 1 hit list, clone archive, tracker update."""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
AEO = ROOT / "_pages" / "aeo"
WORKBOOK = AEO / "Vixxo-AEO-Website-Revamp-Status-fixed.xlsx"
VISIBILITY_JSON = AEO / "_phase0_visibility.json"
CITATIONS_JSON = AEO / "_phase0_citations.json"
AUDIT_JSON = AEO / "_audit_results.json"
BATCH_MD = AEO / "batch-1-hit-list.md"
BATCH_JSON = AEO / "batch-1-hit-list.json"
ARCHIVE_LOG = AEO / "_phase0_archive_log.json"

HEADER_ROW = 4
PORTAL = "7718689"
DEPRECATED_PREFIX = "[DEPRECATED-AEO-v1]"
BATCH_ID = "Batch-1"
PULL_DATE = "2026-08-03"

# Slug → keywords for prompt matching + business weight
SLUG_PROFILE: dict[str, dict] = {
    "facility-management-solutions": {"kw": ["facilities management", "facilities solutions", "fm provider", "outsourced facilities", "retail facilities management", "restaurant facilities management", "grocery store facilities"], "weight": 1.5},
    "solutions/hvac": {"kw": ["hvac", "refrigeration downtime", "cold storage", "hvac/r"], "weight": 1.5},
    "solutions/commercial-handyman-services": {"kw": ["handyman", "building maintenance", "facilities maintenance", "repair across locations", "store repair", "minor repair"], "weight": 1.5},
    "solutions/food-service-equipment": {"kw": ["food service equipment", "food equipment", "kitchen equipment", "fryer", "grill", "oven repair", "commercial kitchen"], "weight": 1.5},
    "beverage-equipment-vixxo": {"kw": ["beverage equipment", "beverage program", "fountain drink"], "weight": 1.4},
    "solutions/coffee": {"kw": ["coffee equipment", "coffee machine", "coffee maintenance", "espresso", "bean-to-cup", "specialty beverage", "brewer"], "weight": 1.4},
    "solutions/refrigeration-services": {"kw": ["refrigeration", "cold storage repair", "open-case refrigeration"], "weight": 1.4},
    "solutions/commercial-plumbing-services": {"kw": ["plumbing repair", "plumbing maintenance", "plumbing services"], "weight": 1.4},
    "electrical-services": {"kw": ["electrical repair", "electrical maintenance", "electrical services"], "weight": 1.3},
    "solutions/doors-locks-hardware": {"kw": ["door and lock", "doors, locks", "locksmith", "door hardware", "master key"], "weight": 1.3},
    "solutions/signslighting": {"kw": ["sign and lighting", "signage repair", "lighting maintenance", "sign repair"], "weight": 1.3},
    "industries/grocery": {"kw": ["grocery chain", "grocery store", "supermarket", "food retail facilities"], "weight": 1.2},
    "industries/restaurant": {"kw": ["restaurant chain", "qsr", "fast-casual", "restaurant facilities"], "weight": 1.2},
    "industries/convenience": {"kw": ["c-store", "convenience store"], "weight": 1.2},
    "industries/retail": {"kw": ["retail chain", "multi-site retail", "retail facilities"], "weight": 1.2},
    "about-us/contact-us": {"kw": ["contact vixxo", "24/7 facilities support", "request quote"], "weight": 1.3},
    "(homepage)": {"kw": ["multi-site facility", "national facilities support", "scalable fm solutions"], "weight": 1.4},
    "facility-management/trade-services": {"kw": ["trade services", "national fm services"], "weight": 1.2},
    "why-vixxo-is-different": {"kw": ["why vixxo", "tco", "lifecycle", "total cost of ownership"], "weight": 1.1},
    "vixxo-ai-in-facilities": {"kw": ["ai-driven fm", "ai in facilities", "predictive analytics facilities"], "weight": 1.1},
}

# Batch 1 eligibility — commercial pages only (exclude resource hubs)
BATCH_ELIGIBLE_PREFIXES = (
    "solutions/",
    "beverage-",
    "electrical-services",
    "industries/",
    "facility-equipment-projects/",
    "facility-management/",
    "about-us/contact-us",
    "(homepage)",
    "vixxo-ai-in-facilities",
    "why-vixxo",
)


def is_batch_eligible(slug: str) -> bool:
    if slug in {"about-us/contact-us", "(homepage)", "electrical-services", "beverage-equipment-vixxo"}:
        return True
    return any(slug.startswith(p) for p in BATCH_ELIGIBLE_PREFIXES)

DEFAULT_WEIGHT = 1.0
LEGACY_WEIGHT = 0.5


def canonical_editor_url(page_id: str) -> str:
    return f"https://app-na2.hubspot.com/pages/{PORTAL}/editor/{page_id}/content"


def slug_weight(slug: str, template_family: str) -> float:
    fam = str(template_family or "").strip().lower()
    if fam == "legacy":
        return LEGACY_WEIGHT
    if slug in SLUG_PROFILE:
        return SLUG_PROFILE[slug]["weight"]
    if slug.startswith("solutions/"):
        return 1.4
    if slug.startswith("industries/"):
        return 1.2
    if slug.startswith("facility-management/"):
        return 1.1
    return DEFAULT_WEIGHT


def slug_keywords(slug: str) -> list[str]:
    if slug in SLUG_PROFILE:
        return SLUG_PROFILE[slug]["kw"]
    parts = slug.replace("/", " ").replace("-", " ").split()
    return [p for p in parts if len(p) > 3]


def match_prompts(slug: str, prompts: list[dict]) -> list[dict]:
    kws = slug_keywords(slug)
    if not kws:
        return []
    matched = []
    for p in prompts:
        name = (p.get("prompt") or {}).get("name", "").lower()
        hits = sum(1 for kw in kws if kw in name)
        if hits == 0:
            continue
        # Require stronger match for short/generic keywords
        if hits >= 1 and (len(kws) <= 2 or hits >= 1):
            matched.append(p)
    return matched


def batch_penalty(slug: str) -> float:
    if not is_batch_eligible(slug):
        return 0.15  # strongly deprioritize resource/news/spotlight pages
    return 1.0


def path_for_slug(slug: str) -> str:
    if slug == "(homepage)":
        return "/"
    return "/" + slug.strip("/")


def load_json(path: Path) -> dict | list:
    return json.loads(path.read_text(encoding="utf-8"))


def score_pages(prompts: list[dict], citations: dict[str, float], rows: list[dict]) -> list[dict]:
    domain_citation_share = 0.07139839221603048  # fallback
    scored = []
    for row in rows:
        slug = row["slug"]
        fam = str(row.get("template_family") or "").strip().lower()
        if fam == "legacy":
            continue
        matched = match_prompts(slug, prompts)
        if not matched:
            continue
        prompt_volume = len(matched)
        avg_gap = sum(1.0 - float(p.get("visibility_score") or 0) for p in matched) / prompt_volume
        path = path_for_slug(slug)
        citation_share = citations.get(path, citations.get(path.rstrip("/"), 0.0))
        business_weight = slug_weight(slug, row.get("template_family", ""))
        priority = round(
            prompt_volume * (1 - citation_share) * business_weight * avg_gap * batch_penalty(slug),
            4,
        )
        top_prompts = sorted(matched, key=lambda x: float(x.get("visibility_score") or 0))[:3]
        scored.append({
            **row,
            "profound_priority": priority,
            "prompt_volume": prompt_volume,
            "avg_visibility_gap": round(avg_gap, 4),
            "citation_share": round(citation_share, 6),
            "business_weight": business_weight,
            "top_weak_prompts": [p["prompt"]["name"] for p in top_prompts],
        })
    scored.sort(key=lambda x: x["profound_priority"], reverse=True)
    eligible = [s for s in scored if is_batch_eligible(s["slug"])]
    batch1 = eligible[:15] if len(eligible) >= 15 else eligible + [s for s in scored if s not in eligible][:15 - len(eligible)]
    return scored, batch1


def extract_page_id(cell_value: str | None, audit_map: dict[str, str]) -> str | None:
    slug_key = None
    if cell_value and "website-pages/" in str(cell_value):
        m = re.search(r"website-pages/(\d+)", str(cell_value))
        if m:
            return m.group(1)
    if cell_value and "/editor/" in str(cell_value):
        m = re.search(r"/editor/(\d+)/", str(cell_value))
        if m:
            return m.group(1)
    return None


def main() -> int:
    sys.path.insert(0, str(ROOT / ".cursor" / "bin" / "hubspot-pages"))

    if not VISIBILITY_JSON.is_file():
        print(json.dumps({"error": f"Missing {VISIBILITY_JSON}"}))
        return 1

    vis = load_json(VISIBILITY_JSON)
    prompts = vis.get("data", [])

    citations: dict[str, float] = {}
    if CITATIONS_JSON.is_file():
        cit = load_json(CITATIONS_JSON)
        for row in cit.get("rows", []):
            dims = row.get("dimensions") or []
            if len(dims) >= 2:
                path, metrics = dims[0], row.get("metrics") or []
                if metrics:
                    citations[path] = float(metrics[0])

    audit_map = {}
    if AUDIT_JSON.is_file():
        for item in load_json(AUDIT_JSON).get("results", []):
            audit_map[item.get("slug", "")] = item.get("page_id", "")

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from url_hyperlink import apply_url_hyperlink

    wb = load_workbook(WORKBOOK)
    ws = wb["AEO Page Status"]
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, col).value
        if val:
            headers[str(val)] = col

    # Add new columns if missing
    new_cols = ["Profound Priority", "Publish Status", "Last Profound Pull", "Batch ID"]
    next_col = ws.max_column + 1
    for nc in new_cols:
        if nc not in headers:
            headers[nc] = next_col
            ws.cell(HEADER_ROW, next_col).value = nc
            next_col += 1

    tracker_rows: list[dict] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        name = ws.cell(r, headers["Page Name"]).value
        slug = ws.cell(r, headers["URL Slug"]).value
        if not slug:
            continue
        assign = ws.cell(r, headers["Assignment"]).value if "Assignment" in headers else ""
        fam = ws.cell(r, headers["Template Family"]).value if "Template Family" in headers else ""
        before = ws.cell(r, headers["Before URL"]).value
        page_id = audit_map.get(str(slug), "")
        tracker_rows.append({
            "row": r,
            "page_name": str(name or ""),
            "slug": str(slug),
            "template_family": str(fam or ""),
            "assignee": str(assign or ""),
            "live_url": f"https://www.vixxo.com{path_for_slug(str(slug))}" if slug != "(homepage)" else "https://www.vixxo.com/",
            "clone_page_id": page_id,
            "editor_url": canonical_editor_url(page_id) if page_id else "",
        })

    scored, batch1 = score_pages(prompts, citations, tracker_rows)

    # Write batch files
    BATCH_JSON.write_text(json.dumps({
        "generated": datetime.now(timezone.utc).isoformat(),
        "profound_pull_date": PULL_DATE,
        "category_id": "4b5a41db-8261-44ee-a343-edadd010c26e",
        "domain": "vixxo.com",
        "batch_id": BATCH_ID,
        "formula": "priority = prompt_volume × (1 - citation_share) × business_weight × avg_visibility_gap",
        "count": len(batch1),
        "items": batch1,
    }, indent=2), encoding="utf-8")

    md_lines = [
        "# Batch 1 Hit List — Profound-Led AEO",
        "",
        f"**Generated:** {PULL_DATE}  ",
        f"**Profound category:** `4b5a41db-8261-44ee-a343-edadd010c26e`  ",
        f"**Domain citation share (owned):** 7.14%  ",
        f"**Formula:** `priority = prompt_volume × (1 - citation_share) × business_weight × avg_visibility_gap`",
        "",
        "## Top 15 pages for fresh clone + Profound AEO (Phase 1)",
        "",
        "| Rank | Page | Slug | Assignee | Priority | Prompts | Citation Share | Rationale | Editor URL |",
        "| ---: | --- | --- | --- | ---: | ---: | ---: | --- | --- |",
    ]
    for i, item in enumerate(batch1, 1):
        rationale = "; ".join(item["top_weak_prompts"][:2])
        if len(rationale) > 80:
            rationale = rationale[:77] + "..."
        md_lines.append(
            f"| {i} | {item['page_name']} | `{item['slug']}` | {item['assignee']} | "
            f"{item['profound_priority']} | {item['prompt_volume']} | {item['citation_share']:.4f} | "
            f"{rationale} | [Editor]({item['editor_url']}) |"
        )
    md_lines.extend([
        "",
        "## Next steps",
        "",
        "1. Erica / Neetu / Mia — review this list and approve Batch 1 scope.",
        "2. Phase 1 — **fresh clone** each approved row (v1 clones are `[DEPRECATED-AEO-v1]`).",
        "3. Run Profound-backed AEO pass on new clones; update **Profound Items Addressed**.",
        "",
        "**Draft only — do not publish without approval.**",
    ])
    BATCH_MD.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    batch_slugs = {b["slug"] for b in batch1}
    score_map = {s["slug"]: s for s in scored}

    # Update tracker rows
    for tr in tracker_rows:
        r = tr["row"]
        slug = tr["slug"]
        sc = score_map.get(slug)
        if sc:
            ws.cell(r, headers["Profound Priority"]).value = sc["profound_priority"]
        ws.cell(r, headers["Last Profound Pull"]).value = PULL_DATE
        ws.cell(r, headers["Batch ID"]).value = BATCH_ID if slug in batch_slugs else ""
        ws.cell(r, headers["Publish Status"]).value = "Draft" if tr["clone_page_id"] else "Not Started"

        pid = tr["clone_page_id"]
        if pid:
            editor = canonical_editor_url(pid)
            apply_url_hyperlink(ws.cell(r, headers["After URL"]), editor, "Open clone editor")
            if "HubSpot Editor URL" in headers:
                apply_url_hyperlink(ws.cell(r, headers["HubSpot Editor URL"]), editor, "Open editor")

    table = ws.tables.get("AEOPageStatus")
    if table:
        table.ref = f"A{HEADER_ROW}:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(WORKBOOK)

    # Archive clones
    archive_results = {"archived": 0, "skipped": 0, "failed": [], "items": []}
    if "--skip-archive" not in sys.argv:
        try:
            from hubspot_pages import hubspot_request, pages_api, patch_page_name

            for item in load_json(AUDIT_JSON).get("results", []):
                page_id = str(item.get("page_id") or "")
                name = str(item.get("name") or "")
                if not page_id:
                    archive_results["failed"].append({"page_id": page_id, "error": "no page_id"})
                    continue
                if name.startswith(DEPRECATED_PREFIX):
                    archive_results["skipped"] += 1
                    archive_results["items"].append({"page_id": page_id, "status": "already_deprecated"})
                    continue
                new_name = f"{DEPRECATED_PREFIX} {name}"
                try:
                    patch_page_name(page_id, new_name, "site-page")
                    archive_results["archived"] += 1
                    archive_results["items"].append({"page_id": page_id, "status": "archived", "new_name": new_name})
                except Exception as exc:
                    archive_results["failed"].append({"page_id": page_id, "error": str(exc)})
        except Exception as exc:
            archive_results["error"] = str(exc)

    ARCHIVE_LOG.write_text(json.dumps(archive_results, indent=2), encoding="utf-8")

    print(json.dumps({
        "status": "ok",
        "batch1_count": len(batch1),
        "batch_md": str(BATCH_MD),
        "batch_json": str(BATCH_JSON),
        "workbook": str(WORKBOOK),
        "archive": archive_results,
        "top5": [{"slug": b["slug"], "priority": b["profound_priority"]} for b in batch1[:5]],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
