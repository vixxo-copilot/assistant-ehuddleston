#!/usr/bin/env python3
"""Inspect Excel hyperlink objects in AEO tracker workbooks."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def inspect(path: Path, rows: list[int] | None = None) -> None:
    rows = rows or [5, 20]
    print(f"\n=== {path.name} ===")
    wb = load_workbook(path, data_only=False)
    ws = wb["AEO Page Status"]
    headers = {
        ws.cell(4, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(4, c).value
    }
    d_col = headers.get("After URL", 4)
    v_col = headers.get("HubSpot Editor URL")
    print(f"After URL col={d_col}, Editor col={v_col}")
    for r in rows:
        for name, col in [("D/After", d_col), ("V/Editor", v_col)]:
            if not col:
                continue
            cell = ws.cell(r, col)
            hl = cell.hyperlink
            val = repr(cell.value)
            if len(val) > 120:
                val = val[:117] + "..."
            print(f"Row {r} {name}: value={val}")
            if hl:
                print(f"  hyperlink.target={getattr(hl, 'target', None)}")
                print(f"  hyperlink.location={getattr(hl, 'location', None)}")
                print(f"  hyperlink.ref={getattr(hl, 'ref', None)}")
                print(f"  hyperlink.display={getattr(hl, 'display', None)}")
                print(f"  hyperlink.id={getattr(hl, 'id', None)}")
            else:
                print("  NO hyperlink object")
            if isinstance(cell.value, str) and cell.value.startswith("="):
                print("  FORMULA cell")


def main() -> int:
    paths = [Path(p) for p in sys.argv[1:]] if len(sys.argv) > 1 else [
        Path(r"c:\Users\EHuddleston\source\assistant-EHuddleston\_pages\aeo\Vixxo-AEO-Website-Revamp-Status.xlsx"),
        Path(r"c:\Users\EHuddleston\source\assistant-EHuddleston\_pages\aeo\Vixxo-AEO-Website-Revamp-Status-upload.xlsx"),
    ]
    for p in paths:
        if p.exists():
            inspect(p)
        else:
            print(f"MISSING: {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
