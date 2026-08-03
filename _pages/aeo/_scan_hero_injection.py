#!/usr/bin/env python3
"""Detect clones with AEO content wrongly injected into hero modules."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / ".cursor" / "bin" / "hubspot-pages"))

import openpyxl  # noqa: E402
from aeo_revamp import _is_hero_rich_text, _iter_rich_text_targets  # noqa: E402
from hubspot_pages import hubspot_request, load_dotenv, pages_api  # noqa: E402

load_dotenv(ROOT)
WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
HEADER_ROW = 4


def extract_page_id(url: str | None) -> str | None:
    if not url:
        return None
    text = str(url)
    if text.upper().startswith("=HYPERLINK("):
        match = re.search(r'HYPERLINK\("([^"]+)"', text, flags=re.I)
        if match:
            text = match.group(1)
    match = re.search(r"/(?:website-pages|landing-pages)/(\d+)/edit", text)
    return match.group(1) if match else None


def hero_has_aeo(page: dict) -> bool:
    for entry in _iter_rich_text_targets(page.get("layoutSections") or {}):
        text = str(entry["obj"].get("rich_text") or "")
        if not _is_hero_rich_text(text, entry.get("module_label") or ""):
            continue
        lower = text.lower()
        if '<h4><span style="color: #8e992e' in lower or "<h4><span style='color: #8e992e" in lower:
            return True
        if any(
            marker in text
            for marker in (
                "Vixxo helps multi-site operators",
                "Vixxo is a national facilities",
                "What is Vixxo",
                "Contact Vixxo to discuss",
                "Vixxo careers connect",
            )
        ):
            return True
    return False


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["AEO Page Status"]
    headers = {c.value: i + 1 for i, c in enumerate(ws[HEADER_ROW]) if c.value}
    api = pages_api("site-page")
    broken: list[dict] = []
    ok: list[str] = []

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        slug = ws.cell(row, headers["URL Slug"]).value
        if not slug:
            continue
        editor = str(ws.cell(row, headers.get("HubSpot Editor URL", 0)).value or "")
        pid = extract_page_id(editor)
        if not pid:
            continue
        try:
            page = hubspot_request("GET", f"{api}/{pid}")
        except SystemExit:
            continue
        if not isinstance(page, dict):
            continue
        if hero_has_aeo(page):
            broken.append({"slug": slug, "page_id": pid, "row": row})
        else:
            ok.append(str(slug))

    print(json.dumps({"broken_count": len(broken), "broken": broken, "ok_count": len(ok)}, indent=2))


if __name__ == "__main__":
    main()
