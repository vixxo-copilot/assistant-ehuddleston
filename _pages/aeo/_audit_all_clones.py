#!/usr/bin/env python3
"""Audit all tracker clone pages for real AEO content optimization."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / ".cursor" / "bin" / "hubspot-pages"))

from hubspot_pages import (  # noqa: E402
    _find_page_by_slug,
    hubspot_request,
    load_dotenv,
    pages_api,
)
from aeo_revamp import _collect_rich_text, verify_clone_content  # noqa: E402

load_dotenv(ROOT)

START = 4
WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"


def extract_page_id(editor_url: str | None) -> str | None:
    if not editor_url:
        return None
    text = str(editor_url)
    if text.upper().startswith("=HYPERLINK("):
        match = re.search(r'HYPERLINK\("([^"]+)"', text, flags=re.I)
        if match:
            text = match.group(1)
    match = re.search(r"/(?:website-pages|landing-pages)/(\d+)/edit", text)
    return match.group(1) if match else None


def audit_page(page: dict) -> dict:
    """Audit clone using same logic as batch verify_clone_content (layout + widgets)."""
    layout_text = _collect_rich_text(page.get("layoutSections") or {})
    widget_text = _collect_rich_text(page.get("widgets") or {})
    container_text = _collect_rich_text(page.get("widgetContainers") or {})
    text = layout_text + widget_text + container_text
    verification = verify_clone_content(page)
    text_len = verification["text_len"]
    has_aeo_injection = verification["has_aeo_injection"]
    has_faq_h3 = text.lower().count("<h3") >= 3 and "?" in text
    meta_desc_len = len(str(page.get("metaDescription") or ""))
    meta_title = page.get("htmlTitle")
    content_confirmed = verification["content_confirmed"]
    meta_only = text_len > 500 and not has_aeo_injection and bool(meta_title or meta_desc_len)
    return {
        "text_len": text_len,
        "layout_text_len": len(layout_text),
        "widget_text_len": len(widget_text) + len(container_text),
        "h1": "<h1" in text.lower(),
        "h2": text.lower().count("<h2"),
        "h3": text.lower().count("<h3"),
        "has_aeo_injection": has_aeo_injection,
        "has_faq_block": has_faq_h3 and has_aeo_injection,
        "meta_title": meta_title,
        "meta_desc_len": meta_desc_len,
        "content_confirmed": content_confirmed,
        "meta_only": meta_only,
        "empty": verification["empty"],
        "body_preview": text[:200].replace("\n", " ").strip(),
    }


def main() -> None:
    import openpyxl

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["AEO Page Status"]
    headers = {c.value: i + 1 for i, c in enumerate(ws[START]) if c.value}
    api = pages_api("site-page")

    results = []
    counts = {"full": 0, "meta_only": 0, "empty": 0, "error": 0, "partial": 0}

    for row in ws.iter_rows(min_row=START + 1):
        slug = row[headers["URL Slug"] - 1].value
        if not slug:
            continue
        name = row[headers["Page Name"] - 1].value
        after_url = str(row[headers["After URL"] - 1].value or "")
        editor_url = str(row[headers["HubSpot Editor URL"] - 1].value or "")
        page_id = extract_page_id(editor_url)

        entry = {
            "slug": slug,
            "name": name,
            "after_url": after_url,
            "editor_url": editor_url,
            "page_id": page_id,
        }

        try:
            if page_id:
                page = hubspot_request("GET", f"{api}/{page_id}")
            else:
                after_slug = after_url.replace("https://www.vixxo.com/", "").rstrip("/")
                found = _find_page_by_slug(after_slug, "site-page") if after_slug else None
                if not found:
                    entry["status"] = "error"
                    entry["error"] = "No page ID or slug match"
                    counts["error"] += 1
                    results.append(entry)
                    continue
                page_id = str(found["id"])
                page = hubspot_request("GET", f"{api}/{page_id}")
                entry["page_id"] = page_id

            audit = audit_page(page)
            entry.update(audit)
            if audit["content_confirmed"]:
                entry["status"] = "full"
                counts["full"] += 1
            elif audit["meta_only"]:
                entry["status"] = "meta_only"
                counts["meta_only"] += 1
            elif audit["empty"]:
                entry["status"] = "empty"
                counts["empty"] += 1
            else:
                entry["status"] = "partial"
                counts["partial"] += 1
        except Exception as exc:  # noqa: BLE001
            entry["status"] = "error"
            entry["error"] = str(exc)[:200]
            counts["error"] += 1

        results.append(entry)

    out = ROOT / "_pages" / "aeo" / "_audit_results.json"
    out.write_text(json.dumps({"counts": counts, "results": results}, indent=2), encoding="utf-8")
    print(json.dumps({"counts": counts, "total": len(results)}, indent=2))
    print("\nMeta-only pages:")
    for r in results:
        if r.get("status") == "meta_only":
            print(f"  {r['slug']} -> {r.get('page_id')} ({r.get('text_len')} chars)")


if __name__ == "__main__":
    main()
