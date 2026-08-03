#!/usr/bin/env python3
"""Extract HubSpot editor URLs from AEO tracker workbook."""
from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
CANDIDATES = [
    ROOT / "Vixxo-AEO-Website-Revamp-Status-fixed.xlsx",
    ROOT / "Vixxo-AEO-Website-Revamp-Status.xlsx",
    ROOT / "Vixxo-AEO-Website-Revamp-Status-upload.xlsx",
]

HYPERLINK_RE = re.compile(r'HYPERLINK\("([^"]*)"', flags=re.I)


def extract_url(cell) -> str:
    val = cell.value
    if val is None:
        return ""
    text = str(val).strip()
    if text.upper().startswith("=HYPERLINK("):
        match = HYPERLINK_RE.search(text)
        if match:
            return match.group(1).replace('""', '"')
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    if text.startswith("http"):
        return text
    return text


def main() -> None:
    xlsx = next((p for p in CANDIDATES if p.exists()), None)
    if not xlsx:
        raise SystemExit("No workbook found")

    wb = load_workbook(xlsx, data_only=False)
    ws = wb.active

    rows: list[tuple[str, str]] = []
    for row in range(5, 99):
        page_name = str(ws.cell(row, 1).value or ws.cell(row, 2).value or "").strip()
        if not page_name:
            continue
        d_url = extract_url(ws.cell(row, 4))
        v_url = extract_url(ws.cell(row, 22)) if ws.max_column >= 22 else ""
        editor_url = d_url if d_url.startswith("http") else (
            v_url if v_url.startswith("http") else d_url or v_url
        )
        rows.append((page_name, editor_url))

    csv_path = ROOT / "all-editor-links.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Page Name", "Editor URL"])
        writer.writerows(rows)

    md_path = ROOT / "ALL-EDITOR-LINKS.md"
    with md_path.open("w", encoding="utf-8") as handle:
        handle.write("# All HubSpot Editor Links — AEO Clone Pages\n\n")
        handle.write(f"Source: `{xlsx.name}` (rows 5–98)\n\n")
        handle.write("## Editor URL pattern\n\n")
        handle.write("```\n")
        handle.write(
            "https://app-na2.hubspot.com/page-ui/{portal_id}/management/pages/"
            "website-pages/{page_id}/edit\n"
        )
        handle.write("```\n\n")
        handle.write(f"**Total pages:** {len(rows)}\n\n")
        handle.write("| Page Name | Editor URL |\n")
        handle.write("| --- | --- |\n")
        for name, url in rows:
            safe_name = name.replace("|", "\\|")
            handle.write(f"| {safe_name} | {url} |\n")

    print(f"Source: {xlsx}")
    print(f"Rows: {len(rows)}")
    missing = [r for r in rows if not r[1].startswith("http")]
    if missing:
        print(f"Missing URLs: {len(missing)}")
        for name, url in missing:
            print(f"  {name!r}: {url!r}")
    print(f"Wrote {csv_path}")
    print(f"Wrote {md_path}")


if __name__ == "__main__":
    main()
