#!/usr/bin/env python3
"""Audit all 94 clones: compare structure vs live published pages."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / ".cursor" / "bin" / "hubspot-pages"))

import openpyxl  # noqa: E402

from aeo_revamp import _collect_rich_text  # noqa: E402
from hubspot_pages import (  # noqa: E402
    _find_page_by_slug,
    hubspot_request,
    load_dotenv,
    pages_api,
)

load_dotenv(ROOT)
api = pages_api("site-page")
WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
HEADER_ROW = 4


def extract_h1(text: str) -> str | None:
    match = re.search(r"<h1[^>]*>(.*?)</h1>", text, re.I | re.S)
    return re.sub(r"<[^>]+>", "", match.group(1)).strip() if match else None


def page_text(page: dict) -> str:
    return (
        _collect_rich_text(page.get("layoutSections") or {})
        + _collect_rich_text(page.get("widgets") or {})
        + _collect_rich_text(page.get("widgetContainers") or {})
    )


def resolve_live(slug: str, before_url: str | None) -> dict | None:
    if slug in {"(homepage)", "", "home"}:
        data = hubspot_request("GET", f"{api}?limit=50&state__in=PUBLISHED_OR_SCHEDULED")
        for item in data.get("results", []) if isinstance(data, dict) else []:
            if not item.get("slug"):
                return item
        return None
    found = _find_page_by_slug(slug, "site-page")
    if found:
        return found
    if before_url:
        slug_from_url = str(before_url).replace("https://www.vixxo.com", "").strip("/")
        if slug_from_url:
            return _find_page_by_slug(slug_from_url, "site-page")
    return None


def compare(live: dict, clone: dict) -> dict:
    live_text = page_text(live)
    clone_text = page_text(clone)
    live_h1 = extract_h1(live_text)
    clone_h1 = extract_h1(clone_text)
    live_layout = len(_collect_rich_text(live.get("layoutSections") or {}))
    clone_layout = len(_collect_rich_text(clone.get("layoutSections") or {}))
    live_widgets = len(
        _collect_rich_text(live.get("widgets") or {})
        + _collect_rich_text(live.get("widgetContainers") or {})
    )
    clone_widgets = len(
        _collect_rich_text(clone.get("widgets") or {})
        + _collect_rich_text(clone.get("widgetContainers") or {})
    )
    live_tpl = live.get("templatePath") or ""
    clone_tpl = clone.get("templatePath") or ""

    issues: list[str] = []
    if not clone_h1:
        issues.append("clone_missing_h1")
    elif live_h1 and clone_h1 and live_h1.lower() != clone_h1.lower():
        issues.append(f"h1_mismatch:{live_h1!r}!={clone_h1!r}")
    if live_tpl != clone_tpl:
        issues.append("template_mismatch")
    if len(clone_text) < 200:
        issues.append("empty_shell")
    elif len(clone_text) < len(live_text) * 0.5:
        issues.append("content_truncated")
    if live_layout > 500 and clone_layout < 200 and live_widgets < 200:
        issues.append("layout_lost")
    if live_widgets > 500 and clone_widgets < 200 and live_layout < 200:
        issues.append("widgets_lost")

    return {
        "live_h1": live_h1,
        "clone_h1": clone_h1,
        "live_text_len": len(live_text),
        "clone_text_len": len(clone_text),
        "live_layout_len": live_layout,
        "clone_layout_len": clone_layout,
        "live_widget_len": live_widgets,
        "clone_widget_len": clone_widgets,
        "template_match": live_tpl == clone_tpl,
        "issues": issues,
        "match_live": len(issues) == 0,
    }


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["AEO Page Status"]
    headers = {c.value: i + 1 for i, c in enumerate(ws[HEADER_ROW]) if c.value}

    results = []
    counts = {"match": 0, "issue": 0, "error": 0}

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        slug = ws.cell(row, headers["URL Slug"]).value
        if not slug:
            continue
        name = ws.cell(row, headers["Page Name"]).value
        before_url = str(ws.cell(row, headers["Before URL"]).value or "")
        editor_url = str(ws.cell(row, headers["HubSpot Editor URL"]).value or "")
        template = ws.cell(row, headers.get("Template Family", 0)).value if "Template Family" in headers else ""
        match = re.search(r"/website-pages/(\d+)/edit", editor_url)
        clone_id = match.group(1) if match else None

        entry: dict = {
            "slug": slug,
            "name": name,
            "template": template,
            "clone_id": clone_id,
        }

        try:
            live = resolve_live(str(slug), before_url)
            if not live:
                entry["status"] = "error"
                entry["error"] = "live_not_found"
                counts["error"] += 1
                results.append(entry)
                continue
            if not clone_id:
                entry["status"] = "error"
                entry["error"] = "no_clone_id"
                counts["error"] += 1
                results.append(entry)
                continue

            live_full = hubspot_request("GET", f"{api}/{live['id']}")
            clone_full = hubspot_request("GET", f"{api}/{clone_id}")
            cmp = compare(live_full, clone_full)
            entry.update(cmp)
            if cmp["match_live"]:
                entry["status"] = "match"
                counts["match"] += 1
            else:
                entry["status"] = "issue"
                counts["issue"] += 1
        except Exception as exc:  # noqa: BLE001
            entry["status"] = "error"
            entry["error"] = str(exc)[:200]
            counts["error"] += 1

        results.append(entry)
        print(f"{entry['status']:6} {slug}: {entry.get('issues', entry.get('error', ''))}")

    out = {
        "counts": counts,
        "total": len(results),
        "results": results,
        "issue_pages": [r for r in results if r.get("status") == "issue"],
    }
    out_path = ROOT / "_pages" / "aeo" / "_live_match_audit.json"
    out_path.write_text(json.dumps(out, indent=2), encoding="utf-8")
    print(json.dumps({"counts": counts, "total": len(results)}, indent=2))


if __name__ == "__main__":
    main()
