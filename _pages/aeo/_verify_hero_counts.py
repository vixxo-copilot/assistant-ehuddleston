#!/usr/bin/env python3
"""Quick hero char-count verify for sample pages."""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / ".cursor" / "bin" / "hubspot-pages"))

from aeo_revamp import _is_hero_rich_text, _iter_rich_text_targets  # noqa: E402
from hubspot_pages import hubspot_request, load_dotenv, pages_api  # noqa: E402

SAMPLE_SLUGS = {
    "beverage-equipment-vixxo": "Beverage Equipment",
    "(homepage)": "Homepage",
    "solutions/hvac": "HVAC",
    "solutions/hvac-services": "HVAC",
}


def hero_len(page: dict) -> int | None:
    for entry in _iter_rich_text_targets(page.get("layoutSections") or {}):
        text = str(entry["obj"].get("rich_text") or "")
        if _is_hero_rich_text(text, entry.get("module_label") or ""):
            return len(text)
    return None


def body_has_aeo(page: dict) -> bool:
    found_hero = False
    for entry in _iter_rich_text_targets(page.get("layoutSections") or {}):
        text = str(entry["obj"].get("rich_text") or "")
        if _is_hero_rich_text(text, entry.get("module_label") or ""):
            found_hero = True
            continue
        if found_hero and (
            "Vixxo helps multi-site operators" in text
            or "Vixxo is a national facilities" in text
            or "Frequently Asked Questions" in text
        ):
            return True
    return False


def extract_page_id(url: str | None) -> str | None:
    import re

    if not url:
        return None
    text = str(url)
    match = re.search(r"/website-pages/(\d+)/edit", text)
    return match.group(1) if match else None


def load_sample_ids() -> dict[str, str]:
    from openpyxl import load_workbook

    wb_path = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["AEO Page Status"]
    headers = {str(ws.cell(4, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(4, c).value}
    found: dict[str, str] = {}
    for row in range(5, ws.max_row + 1):
        slug = str(ws.cell(row, headers["URL Slug"]).value or "").strip()
        label = SAMPLE_SLUGS.get(slug)
        if not label or label in found:
            continue
        editor = str(ws.cell(row, headers["HubSpot Editor URL"]).value or "")
        cid = extract_page_id(editor)
        if cid:
            found[label] = cid
    return found


def main() -> int:
    load_dotenv(ROOT)
    api = pages_api("site-page")
    sample_ids = load_sample_ids()
    sample_ids.setdefault("Beverage Equipment", "367618900671")

    out = []
    for name in ("Beverage Equipment", "HVAC", "Homepage"):
        cid = sample_ids.get(name)
        row: dict = {"page": name, "clone_id": cid}
        if not cid:
            row["error"] = "clone id not found in tracker"
            out.append(row)
            continue
        clone = hubspot_request("GET", f"{api}/{cid}")
        row["clone_hero_len"] = hero_len(clone)
        row["clone_body_aeo"] = body_has_aeo(clone)
        row["editorUrl"] = f"https://app-na2.hubspot.com/page-ui/7718689/management/pages/website-pages/{cid}/edit"
        out.append(row)

    print(json.dumps(out, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
