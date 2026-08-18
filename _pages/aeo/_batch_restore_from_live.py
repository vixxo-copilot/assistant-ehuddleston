#!/usr/bin/env python3
"""Restore all tracker clones from live layout + re-apply AEO in body only (never publish)."""
from __future__ import annotations

import argparse
import copy
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / ".cursor" / "bin" / "hubspot-pages"))

from openpyxl import load_workbook  # noqa: E402

from aeo_revamp import (  # noqa: E402
    _is_hero_rich_text,
    _iter_rich_text_targets,
    backoff_sleep,
    build_revamp_package,
    build_stage_payload,
    verify_clone_content,
)
from hubspot_pages import editor_url, hubspot_request, load_config, load_dotenv, pages_api  # noqa: E402
from run_aeo_revamp_batch import (  # noqa: E402
    HEADER_ROW,
    extract_page_id,
    load_headers,
    resolve_live_page,
    save_workbook,
    list_pages_fn_factory,
)

DEFAULT_WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
DEFAULT_OUTPUT = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status-fixed.xlsx"


def live_hero_text(page: dict[str, Any]) -> str | None:
    for entry in _iter_rich_text_targets(page.get("layoutSections") or {}):
        text = str(entry["obj"].get("rich_text") or "")
        if _is_hero_rich_text(text, entry.get("module_label") or ""):
            return text
    return None


def hero_has_aeo(page: dict[str, Any]) -> bool:
    for entry in _iter_rich_text_targets(page.get("layoutSections") or {}):
        text = str(entry["obj"].get("rich_text") or "")
        if not _is_hero_rich_text(text, entry.get("module_label") or ""):
            continue
        lower = text.lower()
        if '<h4><span style="color: #8e992e' in lower or "<h4><span style='color: #8e992e" in lower:
            return True
        if any(
            marker in text
            for marker in (
                "Vixxo helps multi-site operators",
                "Vixxo is a national facilities",
                "What is Vixxo",
                "Contact Vixxo to discuss",
                "Vixxo careers connect",
            )
        ):
            return True
    return False


def patch_clone_from_live(
    clone_id: str,
    live_page: dict[str, Any],
    package: dict[str, Any],
    cfg: dict[str, Any],
    *,
    clone_slug: str | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    api = pages_api("site-page")
    clone_page = hubspot_request("GET", f"{api}/{clone_id}")
    if not isinstance(clone_page, dict):
        raise SystemExit(f"Could not fetch clone {clone_id}")

    payload = build_stage_payload(
        clone_page,
        package,
        cfg,
        live_page=live_page,
        clone_slug=clone_slug or str(clone_page.get("slug") or ""),
    )

    state = str(clone_page.get("state") or clone_page.get("currentState") or "").upper()
    endpoint = f"{api}/{clone_id}/draft" if state in {"PUBLISHED", "PUBLISHED_OR_SCHEDULED", "SCHEDULED"} else f"{api}/{clone_id}"
    updated = hubspot_request("PATCH", endpoint, payload)
    if not isinstance(updated, dict):
        updated = clone_page

    live_hero = live_hero_text(live_page)
    clone_hero = live_hero_text(updated)
    verification = {
        "hero_matches_live": live_hero == clone_hero if live_hero else True,
        "hero_aeo_injection": hero_has_aeo(updated),
        **verify_clone_content(updated),
    }
    return updated, verification


def main() -> int:
    load_dotenv(ROOT)
    parser = argparse.ArgumentParser(description="Restore tracker clones from live + body-only AEO")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Save workbook copy here")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--skip-slug", action="append", default=[], help="Skip slugs (already fixed)")
    parser.add_argument("--only-slug", action="append", default=[], help="Process only these slugs")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--save-every", type=int, default=10)
    args = parser.parse_args()

    cfg = load_config()
    workbook = args.workbook.expanduser()
    output = args.output.expanduser()
    if not workbook.is_file():
        raise SystemExit(f"Workbook not found: {workbook}")

    if output.resolve() != workbook.resolve() and not output.is_file():
        import shutil

        shutil.copy2(workbook, output)

    wb = load_workbook(output if output.is_file() else workbook)
    ws = wb["AEO Page Status"]
    headers = load_headers(ws)
    skip = {s.lower() for s in args.skip_slug}
    only = {s.lower() for s in args.only_slug}

    started = datetime.now(timezone.utc).isoformat()
    processed: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    hero_broken_before = 0

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        slug = str(ws.cell(row, headers["URL Slug"]).value or "").strip()
        if not slug or slug.lower() in skip:
            continue
        if only and slug.lower() not in only:
            continue
        if args.limit is not None and len(processed) >= args.limit:
            break

        page_name = str(ws.cell(row, headers["Page Name"]).value or "")
        before_url = str(ws.cell(row, headers["Before URL"]).value or "").strip()
        editor_cell = str(ws.cell(row, headers.get("HubSpot Editor URL", 0)).value or "")
        clone_id = extract_page_id(editor_cell)
        if not clone_id:
            failed.append({"row": row, "slug": slug, "error": "No clone page ID in tracker"})
            continue

        try:
            live_page = resolve_live_page(slug, before_url)
            package = build_revamp_package(live_page, list_pages_fn_factory())

            if args.dry_run:
                api = pages_api("site-page")
                clone_page = hubspot_request("GET", f"{api}/{clone_id}")
                if hero_has_aeo(clone_page if isinstance(clone_page, dict) else {}):
                    hero_broken_before += 1
                processed.append({"slug": slug, "status": "dry-run", "clone_id": clone_id})
                continue

            updated, verification = patch_clone_from_live(
                clone_id,
                live_page,
                package,
                cfg,
            )

            status = "success" if verification.get("content_confirmed") and verification.get("hero_matches_live") and not verification.get("hero_aeo_injection") else "partial"
            entry = {
                "row": row,
                "slug": slug,
                "pageName": page_name,
                "clone_id": clone_id,
                "live_id": str(live_page.get("id") or ""),
                "editor_url": editor_url(clone_id, "site-page", cfg),
                "status": status,
                **verification,
            }
            processed.append(entry)
            print(f"[{len(processed)}] {slug}: {status} hero_ok={verification.get('hero_matches_live')} aeo_hero={verification.get('hero_aeo_injection')}")

            now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
            for header, value in {
                "AEO Status": "Draft Ready" if status == "success" else "In Progress",
                "SEO Status": "Draft Ready" if status == "success" else "In Progress",
                "Last Updated": now,
                "Updated By": "batch_restore_from_live",
                "Accomplishments / Notes": (
                    f"Restored layoutSections/widgets from live page {live_page.get('id')}; "
                    f"AEO injected in body module only. Hero match: {verification.get('hero_matches_live')}."
                ),
            }.items():
                col = headers.get(header)
                if col:
                    ws.cell(row, col, value=value)

            if args.save_every and len(processed) % args.save_every == 0:
                save_workbook(ws, output)
        except SystemExit as exc:
            failed.append({"row": row, "slug": slug, "error": str(exc)})
            backoff_sleep(len(failed))
        except Exception as exc:  # noqa: BLE001
            failed.append({"row": row, "slug": slug, "error": str(exc)})
            backoff_sleep(len(failed))

    if not args.dry_run:
        save_workbook(ws, output)

    summary = {
        "startedAt": started,
        "finishedAt": datetime.now(timezone.utc).isoformat(),
        "workbook": str(workbook),
        "output_workbook": str(output),
        "dry_run": args.dry_run,
        "total_processed": len(processed),
        "success": sum(1 for p in processed if p.get("status") == "success"),
        "partial": sum(1 for p in processed if p.get("status") == "partial"),
        "failed": len(failed),
        "hero_broken_before_dry_run": hero_broken_before if args.dry_run else None,
    }
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    log_path = ROOT / "_pages" / "aeo" / f"batch-restore-{stamp}.json"
    log_path.write_text(json.dumps({"summary": summary, "processed": processed, "failed": failed}, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
