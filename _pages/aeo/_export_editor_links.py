"""Export all HubSpot editor links from the AEO tracker workbook."""
import csv
import re
from pathlib import Path

from openpyxl import load_workbook

path = Path(__file__).resolve().parent / "Vixxo-AEO-Website-Revamp-Status-fixed.xlsx"
wb = load_workbook(path, data_only=False)
ws = wb.active


def extract_page_id(cell):
    url = ""
    if cell.hyperlink and cell.hyperlink.target:
        url = cell.hyperlink.target
    value = cell.value
    if isinstance(value, str) and value.upper().startswith("=HYPERLINK"):
        match = re.search(r'HYPERLINK\("([^"]+)"', value, re.I)
        if match:
            url = match.group(1)
        else:
            match = re.search(r"HYPERLINK\('([^']+)'", value, re.I)
            if match:
                url = match.group(1)
    if not url:
        return ""
    match = re.search(r"/(?:website-pages|editor)/(\d+)", url)
    return match.group(1) if match else ""


def editor_content_url(page_id: str) -> str:
    return f"https://app-na2.hubspot.com/pages/7718689/editor/{page_id}/content"


rows = []
for row in range(5, 99):
    name = ws.cell(row, 1).value or ws.cell(row, 2).value or f"Row {row}"
    page_id = extract_page_id(ws.cell(row, 4)) or extract_page_id(ws.cell(row, 22))
    if page_id:
        rows.append((str(name).strip(), editor_content_url(page_id)))

out_dir = path.parent
md_path = out_dir / "ALL-EDITOR-LINKS.md"
csv_path = out_dir / "all-editor-links.csv"

with md_path.open("w", encoding="utf-8") as handle:
    handle.write("# All HubSpot AEO Clone Editor Links (94 pages)\n\n")
    handle.write(
        "Pattern: `https://app-na2.hubspot.com/pages/7718689/editor/{PAGE_ID}/content`\n\n"
    )
    handle.write(
        "Note: `page-ui/.../management/pages/.../edit` opens Page Management UI, not the editor.\n\n"
    )
    handle.write("| Page Name | Editor URL |\n|-----------|------------|\n")
    for name, url in rows:
        handle.write(f"| {name} | {url} |\n")

with csv_path.open("w", encoding="utf-8", newline="") as handle:
    writer = csv.writer(handle)
    writer.writerow(["Page Name", "Editor URL"])
    writer.writerows(rows)

print(f"COUNT={len(rows)}")
for name, url in rows:
    print(f"{name}\t{url}")
