#!/usr/bin/env python3
"""Point After URL (column D) at HubSpot editor links, not live draft slugs."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from url_hyperlink import AFTER_HEADER, EDITOR_HEADER, apply_url_hyperlinks_to_sheet

ROOT = Path(__file__).resolve().parents[3]
DEFAULT_WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
HEADER_ROW = 4
def main() -> int:
    workbook = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    wb = load_workbook(workbook)
    ws = wb["AEO Page Status"]

    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(HEADER_ROW, col).value
        if value:
            headers[str(value)] = col

    after_col = headers.get(AFTER_HEADER)
    editor_col = headers.get(EDITOR_HEADER)
    if not after_col or not editor_col:
        raise SystemExit(f"Missing {AFTER_HEADER!r} or {EDITOR_HEADER!r}")

    table = ws.tables.get("AEOPageStatus")
    link_stats = apply_url_hyperlinks_to_sheet(ws, header_row=HEADER_ROW, table=table)
    updated = link_stats["afterLinked"]
    skipped = link_stats["dataRows"] - updated
    if table is not None:
        table.ref = f"A{HEADER_ROW}:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(workbook)
    print(
        json.dumps(
            {
                "status": "fixed",
                "path": str(workbook),
                "afterColumn": get_column_letter(after_col),
                "editorColumn": get_column_letter(editor_col),
                "rowsUpdated": updated,
                "rowsSkipped": skipped,
                "afterLinked": link_stats["afterLinked"],
                "editorLinked": link_stats.get("editorLinked", 0),
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
