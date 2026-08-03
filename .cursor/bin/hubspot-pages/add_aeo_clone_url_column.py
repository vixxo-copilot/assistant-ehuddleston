#!/usr/bin/env python3
"""Insert After URL column after Before URL in the AEO tracker workbook."""
from __future__ import annotations

import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import TableColumn

ROOT = Path(__file__).resolve().parents[3]
WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
HEADER = "After URL"
INSERT_COL = 4  # D — immediately after Before URL (C)
START_ROW = 4
ASSIGN_HEADER = "Assignment"


def main() -> None:
    wb = load_workbook(WORKBOOK)
    ws = wb["AEO Page Status"]
    table = ws.tables["AEOPageStatus"]

    existing_headers = [ws.cell(START_ROW, c).value for c in range(1, ws.max_column + 1)]
    if HEADER in existing_headers:
        col_idx = existing_headers.index(HEADER) + 1
        ref_col = get_column_letter(col_idx)
        print(
            json.dumps(
                {
                    "status": "already_present",
                    "header": HEADER,
                    "column": ref_col,
                    "tableRef": table.ref,
                    "path": str(WORKBOOK),
                },
                indent=2,
            )
        )
        return

    for merge in list(ws.merged_cells.ranges):
        if str(merge).startswith("A1:") or str(merge).startswith("A2:"):
            ws.unmerge_cells(str(merge))

    ws.insert_cols(INSERT_COL)
    data_end = ws.max_row
    ref_col = get_column_letter(ws.max_column)

    # Header formatting matches adjacent Before URL column.
    style_src = ws.cell(row=START_ROW, column=INSERT_COL - 1)
    header = ws.cell(row=START_ROW, column=INSERT_COL, value=HEADER)
    header.font = copy(style_src.font)
    header.fill = copy(style_src.fill)
    header.alignment = copy(style_src.alignment)
    header.border = copy(style_src.border)

    data_style_src = ws.cell(row=START_ROW + 1, column=INSERT_COL - 1)
    for row_idx in range(START_ROW + 1, data_end + 1):
        cell = ws.cell(row=row_idx, column=INSERT_COL)
        cell.value = ""
        cell.alignment = copy(data_style_src.alignment) or Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions[get_column_letter(INSERT_COL)].width = 36

    # Expand formal Excel table and insert column metadata after Before URL.
    table.ref = f"A{START_ROW}:{ref_col}{data_end}"
    clone_col = TableColumn(id=INSERT_COL, name=HEADER)
    table.tableColumns.insert(INSERT_COL - 1, clone_col)
    for idx, col in enumerate(table.tableColumns, start=1):
        col.id = idx

    ws.merge_cells(f"A1:{ref_col}1")
    ws.merge_cells(f"A2:{ref_col}2")

    wb.save(WORKBOOK)

    assign_col = existing_headers.index(ASSIGN_HEADER) + 1 if ASSIGN_HEADER in existing_headers else None
    if assign_col:
        assign_col += 1  # shifted by insert

    print(
        json.dumps(
            {
                "status": "added",
                "header": HEADER,
                "column": get_column_letter(INSERT_COL),
                "assignmentColumn": get_column_letter(assign_col) if assign_col else None,
                "tableRef": table.ref,
                "tableColumns": [tc.name for tc in table.tableColumns],
                "dataRows": data_end - START_ROW,
                "path": str(WORKBOOK),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
