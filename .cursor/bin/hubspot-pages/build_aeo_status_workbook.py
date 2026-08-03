#!/usr/bin/env python3
"""Build the Vixxo AEO website revamp status Excel workbook."""
from __future__ import annotations

import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

ASSIGNEES = ["Erica Huddleston", "Neetu Rao", "Mia Li"]
ROOT = Path(__file__).resolve().parents[3]
PAGES_SCRIPT = Path(__file__).resolve().parent / "hubspot_pages.py"
DEFAULT_OUTPUT = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"


def ensure_openpyxl():
    try:
        from openpyxl import Workbook  # noqa: F401
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])


def fetch_pages() -> list[dict]:
    proc = subprocess.run(
        [sys.executable, str(PAGES_SCRIPT), "list-pages", "--page-type", "site-page", "--state", "PUBLISHED_OR_SCHEDULED", "--limit", "100"],
        capture_output=True,
        text=True,
        cwd=str(ROOT),
    )
    if proc.returncode != 0:
        raise SystemExit(proc.stderr or proc.stdout)
    data = json.loads(proc.stdout)
    return data.get("pages", [])


def template_family(path: str) -> str:
    p = path or ""
    if "CLEAN-6-1" in p or "clean-pro" in p:
        return "CLEAN-6-1"
    if "generated_layouts" in p:
        return "Legacy"
    return "Other"


def build_workbook(pages: list[dict], output: Path) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    from url_hyperlink import AFTER_HEADER, BEFORE_HEADER, apply_url_hyperlink

    headers = [
        "Page Name",
        "URL Slug",
        BEFORE_HEADER,
        AFTER_HEADER,
        "Template",
        "Template Family",
        "AEO Status",
        "AEO Score (Before)",
        "AEO Score (After)",
        "SEO Status",
        "SEO Score (Before)",
        "SEO Score (After)",
        "Meta Title (After)",
        "Meta Description (After)",
        "Primary Keyword",
        "SEO Notes",
        "Last Updated",
        "Updated By",
        "Profound Items Addressed",
        "Accomplishments / Notes",
        "LLM Test Queries",
        "HubSpot Editor URL",
        "Report File",
        "Assignment",
    ]

    rows = []
    for idx, page in enumerate(sorted(
        pages,
        key=lambda item: (template_family(item.get("templatePath", "")), item.get("slug") or item.get("name") or ""),
    )):
        slug = page.get("slug") or "(homepage)"
        family = template_family(page.get("templatePath", ""))
        status = "Not Started" if family == "CLEAN-6-1" else "Deferred (Legacy Template)"
        note = (
            "Seeded from HubSpot inventory on initial tracker creation"
            if family == "CLEAN-6-1"
            else "Revamp deferred until page is on CLEAN-6-1 template"
        )
        seo_status = status
        rows.append(
            [
                page.get("name") or "",
                slug,
                page.get("url") or "",
                "",
                page.get("templatePath") or "",
                family,
                status,
                "",
                "",
                seo_status,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                note,
                "",
                page.get("editorUrl") or "",
                "",
                ASSIGNEES[idx % len(ASSIGNEES)],
            ]
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "AEO Page Status"

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"] = "Vixxo Website AEO + SEO Revamp Status"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="3E4543")

    seeded = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    ws["A2"] = f"Last seeded: {seeded} | Source: HubSpot CMS site-pages API"

    start_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8E992E")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    before_col = headers.index(BEFORE_HEADER) + 1
    after_col = headers.index(AFTER_HEADER) + 1

    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx in (before_col, after_col) and value and str(value).startswith(("http://", "https://")):
                apply_url_hyperlink(cell, str(value))

    last_row = start_row + len(rows)
    table_ref = f"A{start_row}:{get_column_letter(len(headers))}{last_row}"
    table = Table(displayName="AEOPageStatus", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    widths = [
        34, 28, 36, 36, 42, 14, 16, 12, 12,
        16, 12, 12, 36, 48, 24, 36,
        14, 16, 28, 40, 28, 48, 28, 22,
    ]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = f"A{start_row + 1}"

    summary = wb.create_sheet("Summary")
    summary["A1"] = "AEO + SEO Tracker Summary"
    summary["A1"].font = Font(bold=True, size=14)
    clean_count = sum(1 for row in rows if row[5] == "CLEAN-6-1")
    legacy_count = sum(1 for row in rows if row[5] == "Legacy")
    summary["A3"] = "Total published pages"
    summary["B3"] = len(rows)
    summary["A4"] = "CLEAN-6-1 pages (weekly audit scope)"
    summary["B4"] = clean_count
    summary["A5"] = "Legacy template pages (deferred)"
    summary["B5"] = legacy_count
    summary["A7"] = "SharePoint folder"
    summary["B7"] = "Marketing919 / General / Website & SEO / AEO Website Updates"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return {
        "path": str(output),
        "totalPages": len(rows),
        "clean61": clean_count,
        "legacy": legacy_count,
    }


def main() -> None:
    ensure_openpyxl()
    pages = fetch_pages()
    result = build_workbook(pages, DEFAULT_OUTPUT)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
