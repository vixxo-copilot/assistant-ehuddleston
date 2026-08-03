#!/usr/bin/env python3
"""Insert SEO columns after AEO Score (After) in the AEO+SEO tracker workbook."""
from __future__ import annotations

import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import TableColumn

ROOT = Path(__file__).resolve().parents[3]
WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
START_ROW = 4
INSERT_AFTER = "AEO Score (After)"
SEO_HEADERS = [
    "SEO Status",
    "SEO Score (Before)",
    "SEO Score (After)",
    "Meta Title (After)",
    "Meta Description (After)",
    "Primary Keyword",
    "SEO Notes",
]
SEO_WIDTHS = [16, 12, 12, 36, 48, 24, 36]


def main() -> None:
    wb = load_workbook(WORKBOOK)
    ws = wb["AEO Page Status"]
    table = ws.tables["AEOPageStatus"]

    existing_headers = [ws.cell(START_ROW, c).value for c in range(1, ws.max_column + 1)]
    if all(h in existing_headers for h in SEO_HEADERS):
        print(
            json.dumps(
                {
                    "status": "already_present",
                    "headers": SEO_HEADERS,
                    "tableRef": table.ref,
                    "path": str(WORKBOOK),
                },
                indent=2,
            )
        )
        return

    if INSERT_AFTER not in existing_headers:
        raise SystemExit(f"Anchor column '{INSERT_AFTER}' not found in workbook headers.")

    insert_col = existing_headers.index(INSERT_AFTER) + 2  # 1-based, after anchor
    missing = [h for h in SEO_HEADERS if h not in existing_headers]
    insert_count = len(missing)

    for merge in list(ws.merged_cells.ranges):
        if str(merge).startswith("A1:") or str(merge).startswith("A2:"):
            ws.unmerge_cells(str(merge))

    ws.insert_cols(insert_col, amount=insert_count)
    data_end = ws.max_row

    style_src = ws.cell(row=START_ROW, column=insert_col - 1)
    data_style_src = ws.cell(row=START_ROW + 1, column=insert_col - 1)

    for offset, header in enumerate(missing):
        col_idx = insert_col + offset
        header_cell = ws.cell(row=START_ROW, column=col_idx, value=header)
        header_cell.font = copy(style_src.font)
        header_cell.fill = copy(style_src.fill)
        header_cell.alignment = copy(style_src.alignment)
        header_cell.border = copy(style_src.border)

        for row_idx in range(START_ROW + 1, data_end + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = ""
            cell.alignment = copy(data_style_src.alignment) or Alignment(vertical="top", wrap_text=True)

        width = SEO_WIDTHS[SEO_HEADERS.index(header)]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ref_col = get_column_letter(ws.max_column)
    table.ref = f"A{START_ROW}:{ref_col}{data_end}"

    anchor_idx = existing_headers.index(INSERT_AFTER)
    for offset, header in enumerate(missing):
        tc = TableColumn(id=insert_col + offset, name=header)
        table.tableColumns.insert(anchor_idx + 1 + offset, tc)
    for idx, col in enumerate(table.tableColumns, start=1):
        col.id = idx

    ws.merge_cells(f"A1:{ref_col}1")
    ws.merge_cells(f"A2:{ref_col}2")

    title = ws["A1"].value or ""
    if "SEO" not in str(title):
        ws["A1"] = "Vixxo Website AEO + SEO Revamp Status"

    wb.save(WORKBOOK)

    print(
        json.dumps(
            {
                "status": "added",
                "headers": missing,
                "insertColumn": get_column_letter(insert_col),
                "tableRef": table.ref,
                "tableColumns": [tc.name for tc in table.tableColumns],
                "dataRows": data_end - START_ROW,
                "path": str(WORKBOOK),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
