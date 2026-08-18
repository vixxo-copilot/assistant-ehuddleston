#!/usr/bin/env python3
"""Rename Before/After URL columns and apply live Excel hyperlinks in the AEO tracker.

After URL (column D) hyperlinks prefer HubSpot Editor URL (column V) so clicks
open the clone page editor, not a live slug or the website-pages list.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from url_hyperlink import AFTER_HEADER, BEFORE_HEADER, apply_url_hyperlinks_to_sheet

ROOT = Path(__file__).resolve().parents[3]
WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
START_ROW = 4


def main() -> None:
    output = WORKBOOK
    if len(sys.argv) > 1:
        output = Path(sys.argv[1])

    wb = load_workbook(output)
    ws = wb["AEO Page Status"]
    table = ws.tables.get("AEOPageStatus")

    result = apply_url_hyperlinks_to_sheet(ws, header_row=START_ROW, table=table)

    if table is not None:
        ref_col = get_column_letter(ws.max_column)
        table.ref = f"A{START_ROW}:{ref_col}{ws.max_row}"

    wb.save(output)

    headers = [ws.cell(START_ROW, c).value for c in range(1, ws.max_column + 1)]
    print(
        json.dumps(
            {
                "status": "fixed",
                "path": str(output),
                "beforeHeader": BEFORE_HEADER,
                "afterHeader": AFTER_HEADER,
                "beforeColumn": get_column_letter(result["beforeColumn"]),
                "afterColumn": get_column_letter(result["afterColumn"]),
                "beforeLinked": result["beforeLinked"],
                "afterLinked": result["afterLinked"],
                "editorLinked": result.get("editorLinked", 0),
                "dataRows": result["dataRows"],
                "tableRef": table.ref if table else None,
                "headers": headers,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
