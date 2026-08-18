"""Shared Excel hyperlink helpers for AEO tracker workbooks."""
from __future__ import annotations

import re
from copy import copy

from openpyxl.cell.cell import Cell
from openpyxl.styles import Font

BEFORE_HEADER = "Before URL"
AFTER_HEADER = "After URL"
EDITOR_HEADER = "HubSpot Editor URL"
AFTER_LINK_LABEL = "Open clone editor"
EDITOR_LINK_LABEL = "Open editor"

LEGACY_EDITOR_RE = re.compile(
    r"^https://app(?:-na\d+)?\.hubspot\.com/website/(?P<portal>\d+)/pages/"
    r"(?P<kind>website-pages|landing-pages)/(?P<page_id>\d+)/edit/?$"
)
PAGE_UI_EDITOR_RE = re.compile(
    r"^https://app(?:-na\d+)?\.hubspot\.com/page-ui/(?P<portal>\d+)/management/pages/"
    r"(?P<kind>website-pages|landing-pages)/(?P<page_id>\d+)/edit/?$"
)


def canonical_editor_url(url: str) -> str:
    """Return the page-ui editor URL HubSpot expects (no legacy redirect hop)."""
    text = (url or "").strip()
    if not text:
        return ""
    for pattern in (PAGE_UI_EDITOR_RE, LEGACY_EDITOR_RE):
        match = pattern.match(text)
        if match:
            host = "app-na2.hubspot.com" if "-na2" in text else "app.hubspot.com"
            portal = match.group("portal")
            kind = match.group("kind")
            page_id = match.group("page_id")
            return (
                f"https://{host}/page-ui/{portal}/management/pages/{kind}/{page_id}/edit"
            )
    return text


def is_editor_url(url: str) -> bool:
    """True when url is a HubSpot page editor link."""
    text = (url or "").strip()
    return bool(PAGE_UI_EDITOR_RE.match(text) or LEGACY_EDITOR_RE.match(text))


def _escape_formula_text(value: str) -> str:
    return value.replace('"', '""')


def apply_url_hyperlink(
    cell: Cell,
    url: str,
    display: str | None = None,
    *,
    use_formula: bool = True,
) -> None:
    """Set a clickable hyperlink. Prefer Excel HYPERLINK() for SharePoint Online."""
    text = (display or url or "").strip()
    if not url or not str(url).startswith(("http://", "https://")):
        cell.value = text
        cell.hyperlink = None
        return

    target = canonical_editor_url(str(url)) if is_editor_url(str(url)) else str(url)

    if use_formula:
        safe_url = _escape_formula_text(target)
        safe_text = _escape_formula_text(text)
        cell.value = f'=HYPERLINK("{safe_url}","{safe_text}")'
        cell.hyperlink = None
    else:
        cell.value = text
        cell.hyperlink = target

    base_font = copy(cell.font) if cell.font else Font()
    cell.font = Font(
        name=base_font.name,
        size=base_font.size,
        bold=base_font.bold,
        italic=base_font.italic,
        color="0563C1",
        underline="single",
    )


def apply_url_hyperlinks_to_sheet(
    ws,
    *,
    header_row: int = 4,
    table=None,
) -> dict[str, int]:
    """Apply Before/After URL hyperlinks for all data rows on an AEO tracker sheet."""
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value:
            headers[str(value)] = col

    before_col = headers.get(BEFORE_HEADER)
    after_col = headers.get(AFTER_HEADER)
    editor_col = headers.get(EDITOR_HEADER)
    if not before_col or not after_col:
        raise ValueError(f"Missing {BEFORE_HEADER!r} or {AFTER_HEADER!r} columns")

    before_linked = 0
    after_linked = 0
    editor_linked = 0
    data_rows = 0
    for row in range(header_row + 1, ws.max_row + 1):
        slug_cell = ws.cell(row, headers.get("URL Slug", 2))
        if not slug_cell.value:
            continue
        data_rows += 1
        before_cell = ws.cell(row, before_col)
        after_cell = ws.cell(row, after_col)
        editor_cell = ws.cell(row, editor_col) if editor_col else None
        before_url = str(before_cell.value or "").strip()
        if before_url.startswith("http") and not before_url.startswith("=HYPERLINK("):
            apply_url_hyperlink(before_cell, before_url, before_url, use_formula=True)
            before_linked += 1

        raw_editor = str(editor_cell.value or "").strip() if editor_cell else ""
        if raw_editor.startswith("=HYPERLINK("):
            # Unpack existing formula target if present
            match = re.search(r'=HYPERLINK\("([^"]+)"', raw_editor)
            raw_editor = match.group(1).replace('""', '"') if match else ""
        editor_url = canonical_editor_url(raw_editor)
        if is_editor_url(editor_url):
            if editor_cell is not None:
                editor_cell.value = editor_url
            apply_url_hyperlink(after_cell, editor_url, AFTER_LINK_LABEL, use_formula=True)
            after_linked += 1
            if editor_cell is not None:
                apply_url_hyperlink(editor_cell, editor_url, EDITOR_LINK_LABEL, use_formula=True)
                editor_linked += 1
        else:
            after_url = str(after_cell.value or "").strip()
            if after_url.startswith("=HYPERLINK("):
                match = re.search(r'=HYPERLINK\("([^"]+)"', after_url)
                after_url = match.group(1).replace('""', '"') if match else ""
            after_url = canonical_editor_url(after_url)
            if after_url.startswith(("http://", "https://")):
                apply_url_hyperlink(after_cell, after_url, AFTER_LINK_LABEL, use_formula=True)
                after_linked += 1

    return {
        "beforeColumn": before_col,
        "afterColumn": after_col,
        "editorColumn": editor_col,
        "beforeLinked": before_linked,
        "afterLinked": after_linked,
        "editorLinked": editor_linked,
        "dataRows": data_rows,
    }
