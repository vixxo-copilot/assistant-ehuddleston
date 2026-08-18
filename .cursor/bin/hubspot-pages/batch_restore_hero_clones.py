#!/usr/bin/env python3
"""Restore hero-injected AEO clones from live layout; re-apply AEO to body only."""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from aeo_revamp import (
    _is_hero_rich_text,
    _iter_rich_text_targets,
    backoff_sleep,
    build_revamp_package,
    verify_clone_content,
)
from hubspot_pages import editor_url, hubspot_request, load_config, load_dotenv, pages_api
from run_aeo_revamp_batch import (
    DEFAULT_WORKBOOK,
    HEADER_ROW,
    extract_page_id,
    load_headers,
    patch_clone_page,
    resolve_live_page,
    list_pages_fn_factory,
)

DEFAULT_WORKBOOKS = [
    ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status-fixed.xlsx",
    ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx",
]


def resolve_workbook(path: Path | None) -> Path:
    if path and path.is_file():
        return path
    for candidate in DEFAULT_WORKBOOKS:
        if candidate.is_file():
            return candidate
    raise SystemExit("Tracker workbook not found")


def hero_has_aeo(page: dict) -> bool:
    layout = page.get("layoutSections") or {}
    markers = (
        "Vixxo helps multi-site operators",
        "Frequently Asked Questions",
        "What is Vixxo",
        "Vixxo is a national facilities",
    )
    for entry in _iter_rich_text_targets(layout):
        text = str(entry["obj"].get("rich_text") or "")
        if not _is_hero_rich_text(text, entry.get("module_label") or ""):
            continue
        if len(text) > 250 or any(m in text for m in markers):
            return True
    return False


def count_tracker_rows(ws, headers: dict) -> int:
    total = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if ws.cell(row, headers["URL Slug"]).value:
            total += 1
    return total


def restore_row(
    ws,
    headers: dict,
    row: int,
    *,
    cfg: dict,
    force_all: bool,
) -> dict:
    slug = str(ws.cell(row, headers["URL Slug"]).value or "").strip()
    page_name = str(ws.cell(row, headers["Page Name"]).value or slug)
    before_url = str(ws.cell(row, headers["Before URL"]).value or "").strip()
    editor_cell = (
        str(ws.cell(row, headers["HubSpot Editor URL"]).value or "").strip()
        if "HubSpot Editor URL" in headers
        else ""
    )
    clone_id = extract_page_id(editor_cell)
    if not clone_id:
        return {"status": "skipped", "row": row, "slug": slug, "reason": "no clone id"}

    live_page = resolve_live_page(slug, before_url)
    clone_page = hubspot_request("GET", f"{pages_api('site-page')}/{clone_id}")
    if not isinstance(clone_page, dict):
        raise SystemExit(f"Could not fetch clone {clone_id}")

    if not force_all and not hero_has_aeo(clone_page):
        return {"status": "skipped", "row": row, "slug": slug, "reason": "hero ok"}

    package = build_revamp_package(live_page, list_pages_fn_factory())
    updated, staging_mode = patch_clone_page(
        clone_id,
        live_page,
        package,
        cfg,
        clone_slug=str(clone_page.get("slug") or ""),
    )
    verification = verify_clone_content(updated)
    editor = editor_url(clone_id, "site-page", cfg)
    hero_ok = not hero_has_aeo(updated)
    content_ok = bool(verification.get("content_confirmed"))
    status = "success" if hero_ok and content_ok else "partial"

    return {
        "status": status,
        "row": row,
        "pageName": page_name,
        "slug": slug,
        "clonePageId": clone_id,
        "editorUrl": editor,
        "hero_ok": hero_ok,
        "content_confirmed": content_ok,
        "staging_mode": staging_mode,
        "text_len": verification.get("text_len"),
    }


def main() -> int:
    load_dotenv(ROOT)
    parser = argparse.ArgumentParser(description="Batch restore hero-injected AEO clones")
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--force-all", action="store_true", help="Restore every row with a clone id")
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()

    cfg = load_config()
    workbook = resolve_workbook(args.workbook)
    wb = load_workbook(workbook)
    ws = wb["AEO Page Status"]
    headers = load_headers(ws)
    total = count_tracker_rows(ws, headers)

    started = datetime.now(timezone.utc).isoformat()
    processed: list[dict] = []
    failed: list[dict] = []
    fixed = 0
    idx = 0

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if not ws.cell(row, headers["URL Slug"]).value:
            continue
        idx += 1
        if args.limit is not None and len(processed) >= args.limit:
            break
        slug = ws.cell(row, headers["URL Slug"]).value
        try:
            result = restore_row(ws, headers, row, cfg=cfg, force_all=args.force_all)
            if result.get("status") == "skipped":
                print(f"[{idx}/{total}] SKIP {slug}: {result.get('reason')}")
                continue
            processed.append(result)
            if result.get("status") in ("success", "partial"):
                fixed += 1
            print(
                f"[{idx}/{total}] {result.get('status').upper()} {result.get('pageName')} "
                f"(clone {result.get('clonePageId')}) hero_ok={result.get('hero_ok')}"
            )
        except SystemExit as exc:
            failed.append({"row": row, "slug": slug, "error": str(exc)})
            print(f"[{idx}/{total}] FAIL {slug}: {exc}")
            backoff_sleep(len(failed))
        except Exception as exc:  # noqa: BLE001
            failed.append({"row": row, "slug": slug, "error": str(exc)})
            print(f"[{idx}/{total}] FAIL {slug}: {exc}")
            backoff_sleep(len(failed))

    summary = {
        "fixed": fixed,
        "processed": len(processed),
        "failed": len(failed),
        "total_rows": total,
        "workbook": str(workbook),
    }
    log = {
        "startedAt": started,
        "finishedAt": datetime.now(timezone.utc).isoformat(),
        "summary": summary,
        "processed": processed,
        "failed": failed,
    }
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    log_path = ROOT / "_pages" / "aeo" / f"batch-restore-hero-{stamp}.json"
    log_path.write_text(json.dumps(log, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
