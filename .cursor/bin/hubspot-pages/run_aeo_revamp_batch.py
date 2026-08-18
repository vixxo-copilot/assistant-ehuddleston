#!/usr/bin/env python3
"""Batch AEO + SEO revamp for Vixxo tracker rows (clone drafts only — never publish)."""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from aeo_revamp import (
    backoff_sleep,
    build_revamp_package,
    build_stage_payload,
    layout_has_content,
    merge_structure_payload,
    score_after,
    score_before,
    verify_clone_content,
    write_report_md,
)
from hubspot_pages import (
    _find_page_by_slug,
    _page_summary,
    clone_page_title,
    editor_url,
    hubspot_request,
    load_config,
    load_dotenv,
    pages_api,
    patch_page_name,
    resolve_clone_base_name,
)
from url_hyperlink import AFTER_HEADER, EDITOR_HEADER, apply_url_hyperlink, apply_url_hyperlinks_to_sheet

DEFAULT_WORKBOOK = ROOT / "_pages" / "aeo" / "Vixxo-AEO-Website-Revamp-Status.xlsx"
REPORTS_DIR = ROOT / "_pages" / "aeo" / "reports"
HEADER_ROW = 4
PROFOUND_ITEMS = (
    "Answer-first intro; FAQ block; question-style H2s; meta title/description; "
    "internal links; heading hierarchy per Profound analysis"
)


def slug_to_report_path(slug: str) -> Path:
    safe = (slug or "homepage").replace("/", "-").replace("(", "").replace(")", "")
    if safe in {"", "homepage"}:
        safe = "homepage"
    return REPORTS_DIR / f"{safe}.md"


def extract_page_id(url: str | None) -> str | None:
    if not url:
        return None
    text = str(url)
    if text.upper().startswith("=HYPERLINK("):
        match = re.search(r'HYPERLINK\("([^"]+)"', text, flags=re.I)
        if match:
            text = match.group(1)
    match = re.search(r"/website-pages/(\d+)/edit", text)
    return match.group(1) if match else None


def load_headers(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(HEADER_ROW, col).value
        if value:
            headers[str(value)] = col
    return headers


def list_pages_fn_factory(page_type: str = "site-page"):
    api = pages_api(page_type)

    def _list() -> list[dict[str, Any]]:
        data = hubspot_request("GET", f"{api}?limit=100&state__in=PUBLISHED_OR_SCHEDULED")
        return data.get("results", []) if isinstance(data, dict) else []

    return _list


def clone_live_page(
    source_id: str,
    page_type: str = "site-page",
    *,
    live_page: dict[str, Any] | None = None,
) -> dict[str, Any]:
    api = pages_api(page_type)
    cfg = load_config()
    cloned = hubspot_request("POST", f"{api}/{source_id}/clone", {})
    if not isinstance(cloned, dict):
        raise SystemExit(f"Unexpected clone response for page {source_id}")
    page_id = str(cloned.get("id") or "")
    base_name = resolve_clone_base_name(live_page=live_page, clone_page=cloned)
    dated_name = clone_page_title(base_name)
    if dated_name != str(cloned.get("name") or ""):
        updated = patch_page_name(page_id, dated_name, page_type)
        if isinstance(updated, dict):
            cloned = updated
    summary = _page_summary(cloned, page_type)
    summary["editorUrl"] = editor_url(page_id, page_type, cfg) if page_id else None
    return summary


def resolve_live_page(slug: str, before_url: str | None, page_type: str = "site-page") -> dict[str, Any]:
    api = pages_api(page_type)
    if slug in {"(homepage)", "", "home"}:
        data = hubspot_request("GET", f"{api}?limit=100&state__in=PUBLISHED_OR_SCHEDULED")
        for item in data.get("results", []) if isinstance(data, dict) else []:
            if not item.get("slug"):
                return item
        for item in data.get("results", []) if isinstance(data, dict) else []:
            url = str(item.get("url") or "").rstrip("/")
            if url.endswith("vixxo.com"):
                return item
    lookup_slug = slug if slug not in {"(homepage)", "", "home"} else None
    if lookup_slug:
        found = _find_page_by_slug(lookup_slug, page_type)
        if found:
            return found
    if before_url:
        slug_from_url = str(before_url).replace("https://www.vixxo.com", "").strip("/")
        if slug_from_url:
            found = _find_page_by_slug(slug_from_url, page_type)
            if found:
                return found
    raise SystemExit(f"Could not resolve live page for slug {slug!r}")


def patch_clone_page(
    clone_id: str,
    live_page: dict[str, Any],
    package: dict[str, Any],
    cfg: dict[str, Any],
    *,
    clone_slug: str | None = None,
    page_type: str = "site-page",
) -> tuple[dict[str, Any], str]:
    api = pages_api(page_type)
    clone_page = hubspot_request("GET", f"{api}/{clone_id}")
    if not isinstance(clone_page, dict):
        raise SystemExit(f"Could not fetch clone {clone_id}")

    payload = (
        merge_structure_payload(live_page, clone_page, package)
        if live_page and layout_has_content(live_page)
        else build_stage_payload(
            clone_page,
            package,
            cfg,
            live_page=live_page,
            clone_slug=clone_slug or str(clone_page.get("slug") or ""),
        )
    )
    staging_mode = "patch" if any(k in payload for k in ("layoutSections", "widgets", "widgetContainers")) else "full"

    state = str(clone_page.get("state") or clone_page.get("currentState") or "").upper()
    if state in {"PUBLISHED", "PUBLISHED_OR_SCHEDULED", "SCHEDULED"}:
        updated = hubspot_request("PATCH", f"{api}/{clone_id}/draft", payload)
    else:
        updated = hubspot_request("PATCH", f"{api}/{clone_id}", payload)

    if not isinstance(updated, dict):
        updated = clone_page
    return updated, staging_mode


def update_tracker_row(
    ws,
    headers: dict[str, int],
    row: int,
    *,
    package: dict[str, Any],
    before_scores: tuple[int, int],
    after_scores: tuple[int, int],
    clone_summary: dict[str, Any],
    report_path: Path,
    content_confirmed: bool,
    notes: str,
    updated_by: str = "run_aeo_revamp_batch",
) -> None:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    status = "Draft Ready" if content_confirmed else "In Progress"
    editor = str(clone_summary.get("editorUrl") or "")

    mapping = {
        "After URL": editor,
        "HubSpot Editor URL": editor,
        "AEO Status": status,
        "SEO Status": status,
        "AEO Score (Before)": before_scores[0],
        "AEO Score (After)": after_scores[0],
        "SEO Score (Before)": before_scores[1],
        "SEO Score (After)": after_scores[1],
        "Meta Title (After)": package.get("htmlTitle"),
        "Meta Description (After)": package.get("metaDescription"),
        "Primary Keyword": package.get("primaryKeyword"),
        "SEO Notes": package.get("seoNotes"),
        "LLM Test Queries": package.get("llmTestQueries"),
        "Profound Items Addressed": PROFOUND_ITEMS if content_confirmed else "",
        "Last Updated": now,
        "Updated By": updated_by,
        "Accomplishments / Notes": notes,
        "Report File": str(report_path.relative_to(ROOT)).replace("\\", "/"),
    }

    after_col = headers.get(AFTER_HEADER)
    for header, value in mapping.items():
        col = headers.get(header)
        if not col:
            continue
        cell = ws.cell(row, col, value=value)
        if header in (AFTER_HEADER, EDITOR_HEADER) and editor.startswith("https://"):
            apply_url_hyperlink(cell, editor, editor)


def process_row(
    ws,
    headers: dict[str, int],
    row: int,
    *,
    cfg: dict[str, Any],
    dry_run: bool,
    reuse_clone: bool,
    reprocess_empty: bool,
    reprocess_all: bool,
    assignment: str | None,
) -> dict[str, Any]:
    slug = str(ws.cell(row, headers["URL Slug"]).value or "").strip()
    page_name = str(ws.cell(row, headers["Page Name"]).value or "")
    before_url = str(ws.cell(row, headers["Before URL"]).value or "").strip()
    template_family = str(ws.cell(row, headers["Template Family"]).value or "") if "Template Family" in headers else ""
    assignee = str(ws.cell(row, headers["Assignment"]).value or "").strip() if "Assignment" in headers else ""
    editor_url_cell = str(ws.cell(row, headers["HubSpot Editor URL"]).value or "").strip() if "HubSpot Editor URL" in headers else ""
    aeo_status = str(ws.cell(row, headers["AEO Status"]).value or "") if "AEO Status" in headers else ""

    if assignment and assignee and assignment.lower() not in assignee.lower():
        return {"status": "skipped", "row": row, "slug": slug, "notes": f"Assignment filter ({assignment})"}

    if reprocess_empty and not reprocess_all:
        existing_id = extract_page_id(editor_url_cell)
        if existing_id:
            try:
                existing = hubspot_request("GET", f"{pages_api('site-page')}/{existing_id}")
                if verify_clone_content(existing).get("content_confirmed"):
                    return {
                        "status": "skipped",
                        "row": row,
                        "slug": slug,
                        "notes": "Clone already has verified content",
                    }
            except SystemExit:
                pass
        elif aeo_status == "Draft Ready":
            return {"status": "skipped", "row": row, "slug": slug, "notes": "Already Draft Ready"}

    if dry_run:
        live_page = resolve_live_page(slug, before_url)
        package = build_revamp_package(live_page, list_pages_fn_factory())
        before_scores = score_before(live_page)
        after_scores = score_after(package)
        return {
            "status": "dry-run",
            "row": row,
            "pageName": page_name,
            "slug": slug,
            "sourcePageId": str(live_page.get("id") or ""),
            "beforeUrl": before_url or live_page.get("url"),
            "aeoBefore": before_scores[0],
            "aeoAfter": after_scores[0],
            "seoBefore": before_scores[1],
            "seoAfter": after_scores[1],
            "metaTitle": package.get("htmlTitle"),
            "metaDescription": package.get("metaDescription"),
            "primaryKeyword": package.get("primaryKeyword"),
            "llmTestQueries": package.get("llmTestQueries"),
            "seoNotes": package.get("seoNotes"),
            "notes": "Dry-run validation only — no clone created.",
        }

    live_page = resolve_live_page(slug, before_url)
    source_id = str(live_page.get("id") or "")
    list_pages = list_pages_fn_factory()
    package = build_revamp_package(live_page, list_pages)
    before_scores = score_before(live_page)
    after_scores = score_after(package)

    clone_id = extract_page_id(editor_url_cell) if reuse_clone else None
    if clone_id:
        clone_summary = _page_summary(
            hubspot_request("GET", f"{pages_api('site-page')}/{clone_id}"),
            "site-page",
        )
        clone_summary["editorUrl"] = editor_url_cell or editor_url(clone_id, "site-page", cfg)
    else:
        clone_summary = clone_live_page(source_id, live_page=live_page)
        clone_id = str(clone_summary.get("id") or "")

    clone_page, staging_mode = patch_clone_page(
        clone_id,
        live_page,
        package,
        cfg,
        clone_slug=str(clone_summary.get("slug") or ""),
    )
    clone_summary = _page_summary(clone_page, "site-page")
    clone_summary["editorUrl"] = editor_url(clone_id, "site-page", cfg)

    verification = verify_clone_content(clone_page)
    content_confirmed = bool(verification.get("content_confirmed"))
    report_path = slug_to_report_path(slug)
    write_report_md(
        report_path,
        live_page=live_page,
        clone_page=clone_summary,
        package=package,
        before_scores=before_scores,
        after_scores=after_scores,
    )

    status = "success" if content_confirmed else "partial"
    notes = (
        f"Clone draft staged ({staging_mode}) with AEO answer-first, FAQ, internal links, and SEO meta. "
        f"Clone ID {clone_id}. Content verified: {content_confirmed}."
    )
    update_tracker_row(
        ws,
        headers,
        row,
        package=package,
        before_scores=before_scores,
        after_scores=after_scores,
        clone_summary=clone_summary,
        report_path=report_path,
        content_confirmed=content_confirmed,
        notes=notes,
    )

    return {
        "status": status,
        "row": row,
        "pageName": page_name,
        "slug": slug,
        "sourcePageId": source_id,
        "clonePageId": clone_id,
        "beforeUrl": before_url or live_page.get("url"),
        "afterUrl": clone_summary.get("editorUrl"),
        "plannedLiveUrl": clone_summary.get("url"),
        "editorUrl": clone_summary.get("editorUrl"),
        "aeoStatus": "Draft Ready" if content_confirmed else "In Progress",
        "seoStatus": "Draft Ready" if content_confirmed else "In Progress",
        "aeoBefore": before_scores[0],
        "aeoAfter": after_scores[0],
        "seoBefore": before_scores[1],
        "seoAfter": after_scores[1],
        "metaTitle": package.get("htmlTitle"),
        "metaDescription": package.get("metaDescription"),
        "primaryKeyword": package.get("primaryKeyword"),
        "llmTestQueries": package.get("llmTestQueries"),
        "seoNotes": package.get("seoNotes"),
        "reportFile": str(report_path.relative_to(ROOT)).replace("\\", "/"),
        "content_confirmed": content_confirmed,
        "text_len": verification.get("text_len"),
        "staging_mode": staging_mode,
        "templateFamily": template_family,
        "notes": notes,
    }


def save_workbook(ws, workbook: Path) -> None:
    table = ws.tables.get("AEOPageStatus")
    if table is not None:
        table.ref = f"A{HEADER_ROW}:{get_column_letter(ws.max_column)}{ws.max_row}"
    apply_url_hyperlinks_to_sheet(ws, header_row=HEADER_ROW, table=table)
    ws.parent.save(workbook)


def main() -> int:
    load_dotenv(ROOT)
    parser = argparse.ArgumentParser(description="Batch AEO+SEO revamp for tracker rows")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--reuse-clone", action="store_true", help="Reuse existing clone from editor URL column")
    parser.add_argument("--reprocess-empty", action="store_true", help="Only rows without verified clone content")
    parser.add_argument(
        "--reprocess-all",
        action="store_true",
        help="Force reprocess every row (restore live layout + re-apply AEO)",
    )
    parser.add_argument("--save-every", type=int, default=10, help="Save workbook every N processed rows")
    parser.add_argument(
        "--skip-tracker-save",
        action="store_true",
        help="Do not write the Excel tracker (HubSpot updates only)",
    )
    parser.add_argument("--assignment", help="Filter rows by Assignment column (substring match)")
    args = parser.parse_args()

    cfg = load_config()
    workbook = args.workbook.expanduser()
    if not workbook.is_file():
        raise SystemExit(f"Workbook not found: {workbook}")

    wb = load_workbook(workbook)
    ws = wb["AEO Page Status"]
    headers = load_headers(ws)

    started = datetime.now(timezone.utc).isoformat()
    processed: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    total_rows = 0

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if not ws.cell(row, headers["URL Slug"]).value:
            continue
        if args.limit is not None and len(processed) >= args.limit:
            break
        total_rows += 1
        slug = ws.cell(row, headers["URL Slug"]).value
        try:
            result = process_row(
                ws,
                headers,
                row,
                cfg=cfg,
                dry_run=args.dry_run,
                reuse_clone=args.reuse_clone,
                reprocess_empty=args.reprocess_empty,
                reprocess_all=args.reprocess_all,
                assignment=args.assignment,
            )
            if result.get("status") == "skipped":
                continue
            processed.append(result)
            if result.get("status") == "success":
                print(f"Progress: {sum(1 for p in processed if p.get('status') in ('success', 'partial'))}/{total_rows} rows")
            if args.save_every and len(processed) % args.save_every == 0 and not args.dry_run and not args.skip_tracker_save:
                save_workbook(ws, workbook)
        except SystemExit as exc:
            failed.append({"row": row, "slug": slug, "error": str(exc)})
            backoff_sleep(len(failed))
        except Exception as exc:  # noqa: BLE001
            failed.append({"row": row, "slug": slug, "error": str(exc)})
            backoff_sleep(len(failed))

    if not args.dry_run and not args.skip_tracker_save:
        save_workbook(ws, workbook)

    finished = datetime.now(timezone.utc).isoformat()
    summary = {
        "total": len(processed),
        "succeeded": sum(1 for p in processed if p.get("status") == "success"),
        "failed": len(failed),
        "workbook": str(workbook),
    }
    log = {
        "startedAt": started,
        "workbook": str(workbook),
        "dryRun": args.dry_run,
        "limit": args.limit,
        "processed": processed,
        "failed": failed,
        "summary": summary,
        "finishedAt": finished,
    }
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    log_path = ROOT / "_pages" / "aeo" / f"batch-run-{stamp}.json"
    log_path.write_text(json.dumps(log, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
